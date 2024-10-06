import random

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

dataframe = openpyxl.load_workbook("sorted links.xlsx")
print("workbook loaded")

dataframe1 = dataframe.active
links_list = []
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1):
        links = col[row].value
        links_list.append(links)

def get_random_user_agent():
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36",
        "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:89.0) Gecko/20100101 Firefox/89.0",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36",
        "Mozilla/5.0 (Linux; Android 10; Pixel 3 XL Build/QP1A.190711.020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Mobile Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
        "Mozilla/5.0 (iPad; CPU OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
        "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:60.0) Gecko/20100101 Firefox/60.0"
    ]
    return random.choice(user_agents)


chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("detach", True)

user_agent = get_random_user_agent()
chrome_options.add_argument(f"user-agent={user_agent}")

driver = webdriver.Chrome(options=chrome_options)

maxLink = dataframe1.max_row +1
currentLinkIndex = 704
while currentLinkIndex == maxLink:

    driver.get(f"{links_list[currentLinkIndex]}")

    time.sleep(2)

    companyName = ""
    category = ""
    category_list = []

    try:

        html_list = driver.find_element(BY.XPATH, value="//*[@id='finder-products']/div/div/section/div[2]/div/div/div[1]/ul")
        items = html_list.find_elements_by_tag_name("li")
        for item in items:
        
            text = item.text
            category_list.append(text)
            category = ", ".joined(category_list)
            print(category)


        # name = driver.find_element(By.XPATH, value="//*[@id='profile-title']/h1")
        # name_text = name.text
    except:
        pass

    try:
        companyName = driver.find_element(By.XPATH, value="//*[@id='profile-business-data']/div[1]/div[1]/div/div[2]")
        name = companyName.text
    except:
        pass


    driver.execute_script("window.open('https://docs.google.com/forms/d/1AU00wx_0nrh6wVZi5IoRivx5RKCnOPIp6tQ9GzaZloA/viewform?chromeless=1&edit_requested=true', '_blank');")
    wait = WebDriverWait(driver, 10)


    print("form accessed")
    # Switch to the Google Form tab
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(4)

    print("form loaded")

    company_name = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div[2]/textarea")
    company_name.send_keys(name)

    categories = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div[2]/textarea")
    categories.send_keys(category)

    submit = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='mG61Hd']/div[2]/div/div[3]/div[1]/div[1]/div")))

    submit.send_keys(Keys.ENTER)

    driver.close()

    driver.switch_to.window(driver.window_handles[-1])
    print(currentLinkIndex)

    currentLinkIndex += 1
   
    if currentLinkIndex % 15 == 0:
        print("break time")
        time.sleep(20)

